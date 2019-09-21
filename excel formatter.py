"""stores a list of PS objects. """

from openpyxl import Workbook
from openpyxl import load_workbook
from PS_mapping import PS_PK,PS_SURV,PS_ZONE,PS_STREET,PS_SECTION,PS_SIDE,PS_RESTR,PS_LAT,PS_LONG,PS_PLATES_START
from PS_class import PS
import datetime
import time
import sys
import os.path


def get_data():

    def get_unique_c(plate_list):
        unique_list = []
        if plate_list[0] == None:
            unique_list += [0]
        else:
            unique_list += [1]
        for i in range(1,len(plate_list)):
            if plate_list[i] == None:
                unique_list += [0]
            elif plate_list[i] == plate_list[i-1]:
                unique_list += [0]
            else:
                unique_list += [1]
        return unique_list

    def get_count(plate_list):
        count_list = []
        for plate in plate_list:
            if plate == None:
                count_list += [0]
            else:
                count_list += [1]
        return count_list
    
    while True:
        file_name = input("Enter .xlsx file name: ") + ".xlsx"
        try:
            workbook = load_workbook(filename=file_name)
        except (FileNotFoundError):
            print("File not found in this folder\nMake sure file is in the same folder as this script\n")
        else:
            break
    sheets_list = workbook.sheetnames
    
    while True:
        print("\nChoose a sheet number from this workbook:\n")
        for i in range(len(sheets_list)):
            print("[{}. {}]".format(i+1,sheets_list[i]),end='  ')
            
        sheet_index = input("\nSheet number: ")
        try:
            workbook.active = int(sheet_index)-1
        except ValueError:
            print("Sheet number invalid")
        else:
            print("Loading data from '{}'".format(sheets_list[int(sheet_index)-1]))
            break
        
    sheet = workbook.active

    parking_spaces = []
    dates = []
    times = []
    types = []
    for date in sheet.iter_rows(max_row=1, min_col= PS_PLATES_START +1, values_only=True):
        dates += [date]
    for time in sheet.iter_rows(min_row=2, max_row=2, min_col= PS_PLATES_START +1, values_only=True):
        times += [time]
    for day in dates[0]:
        split_date = day.split("/")
        if datetime.date(int(split_date[2]),int(split_date[1]),int(split_date[0])).weekday()<5:
            types += ["Weekday"]
        else:
            types += ["Weekend"]
            
    for row in sheet.iter_rows(min_row=3, values_only=True):
        parking_space = PS(pk = row[PS_PK],
            Surveyor = row[PS_SURV],
            Zone = row[PS_ZONE],
            Type = types,
            Street = row[PS_STREET],
            Section = row[PS_SECTION],
            Side = row[PS_SIDE],
            Restriction = row[PS_RESTR],
            LatLong = (row[PS_LAT],row[PS_LONG]),
            Plates = row[PS_PLATES_START:],
            UniqueCount = get_unique_c(row[PS_PLATES_START:]),
            Count = get_count(row[PS_PLATES_START:]),
            Times = times,
            Dates = dates)
        parking_spaces.append(parking_space)
    print("Data has been loaded into Python objects\n")
    print("----------------------------------------------------------------------------------")
    print(parking_spaces[0])
    print("----------------------------------------------------------------------------------")
    print("Please check the above object for correctness before proceeding")
    return parking_spaces

def format_data(data):

    def overwrite_output():
        while True:
            overwrite = input("File already exists. Are you sure you want to overwrite? (y/n)\n")
            if overwrite == "n":
                return False
            elif overwrite == "y":
                return True
            else:
                print("Please use 'y' for yes or 'n' for no\n")
        
    while True:
        output_file = input("\nEnter output .xlsx file name: ")
        if os.path.isfile(output_file+".xlsx"):
            if overwrite_output():
                break
        else:
            break
    
    start = time.time()
    pks_in_data = len(data)
    runs_in_data = len(data[0].Plates)

    new_workbook = Workbook()
    sheet = new_workbook.active

    column_names = ["PK","Surveyor","Zone","Date","Type","Street","Section","Side",\
    "Restriction","Latitude", "Longitude", "Time", "Plate", "Count", \
    "Unique Count", "Total Spaces"]
    for i in range(len(column_names)):
        sheet.cell(row=1,column=i+1).value = column_names[i]
    expected_rows = pks_in_data*runs_in_data
    current_row = 2
    for run in range(runs_in_data):
        for ps in data:
            sheet.cell(row=current_row,column=1).value = ps.pk
            sheet.cell(row=current_row,column=2).value = ps.Surveyor
            sheet.cell(row=current_row,column=3).value = ps.Zone
            sheet.cell(row=current_row,column=4).value = ps.Dates[0][run]
            sheet.cell(row=current_row,column=5).value = ps.Type[run]
            sheet.cell(row=current_row,column=6).value = ps.Street
            sheet.cell(row=current_row,column=7).value = ps.Section
            sheet.cell(row=current_row,column=8).value = ps.Side
            sheet.cell(row=current_row,column=9).value = ps.Restriction
            sheet.cell(row=current_row,column=10).value = ps.LatLong[0]
            sheet.cell(row=current_row,column=11).value = ps.LatLong[1]
            sheet.cell(row=current_row,column=12).value = ps.Times[0][run]
            if isinstance(ps.Plates[run],str):
                sheet.cell(row=current_row,column=13).value = ps.Plates[run]
            sheet.cell(row=current_row,column=14).value = ps.Count[run]
            sheet.cell(row=current_row,column=15).value = ps.UniqueCount[run]
            sheet.cell(row=current_row,column=16).value = 1
            print("{}/{} rows formatted. {}%".format(current_row-1,expected_rows,round((current_row-1)/expected_rows*100),5), flush=True)
            sys.stdout.flush()
            current_row += 1
            

    new_workbook.save(filename=output_file+".xlsx")
    print("------------------------------------------------------")
    print("Data has been saved in output file \"{}.xlsx\".\n\
Expected rows = {}*{}+1 = {} rows total".format(output_file,pks_in_data,runs_in_data,expected_rows+1))
    end= time.time()
    print("This operation took {} seconds\n".format(round(end-start,1)))
    input("Press enter to exit\n")


data = get_data()
format_data(data)

